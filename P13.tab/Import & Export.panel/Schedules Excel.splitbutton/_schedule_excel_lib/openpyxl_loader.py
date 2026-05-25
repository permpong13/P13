# -*- coding: utf-8 -*-
from __future__ import print_function

import json
import os
import subprocess
import sys
import tempfile

def _get_common_paths():
    paths = [
        r"C:\Python313\python.exe",
        r"C:\Python312\python.exe",
        r"C:\Python311\python.exe",
        r"C:\Python310\python.exe",
        r"C:\Python39\python.exe",
        r"C:\Program Files\Python313\python.exe",
        r"C:\Program Files\Python312\python.exe",
        r"C:\Program Files\Python311\python.exe",
    ]
    userprofile = os.environ.get("USERPROFILE") or os.environ.get("HOMEPATH") or os.path.expanduser("~")
    if userprofile:
        for v in ["313", "312", "311", "310", "39"]:
            paths.append(os.path.join(userprofile, r"AppData\Local\Programs\Python\Python" + v, "python.exe"))
    return paths


def _decode_output(value):
    try:
        return value.decode("utf-8", "ignore")
    except Exception:
        return str(value)


def find_cpython():
    try:
        command = ["where", "python"] if sys.platform == "win32" else ["which", "python3"]
        output = subprocess.check_output(command, stderr=subprocess.STDOUT)
        for line in _decode_output(output).strip().splitlines():
            candidate = line.strip()
            if not candidate or not os.path.isfile(candidate) or "WindowsApps" in candidate:
                continue
            test = subprocess.check_output(
                [candidate, "-c", "import sys; print(sys.implementation.name)"],
                stderr=subprocess.STDOUT
            )
            if _decode_output(test).strip() == "cpython":
                return candidate
    except Exception:
        pass

    for path in _get_common_paths():
        if os.path.isfile(path):
            return path

    python_home = os.environ.get("PYTHON_HOME", "")
    if python_home:
        candidate = os.path.join(python_home, "python.exe")
        if os.path.isfile(candidate):
            return candidate

    return None


def get_cpython_site_packages(python_exe):
    try:
        output = subprocess.check_output(
            [python_exe, "-c", "import site; print('\\n'.join(site.getsitepackages()))"],
            stderr=subprocess.STDOUT
        )
        return [path.strip() for path in _decode_output(output).splitlines() if path.strip()]
    except Exception:
        return []


def ensure_openpyxl_installed(python_exe):
    try:
        subprocess.check_output([python_exe, "-c", "import openpyxl"], stderr=subprocess.STDOUT)
        return True
    except subprocess.CalledProcessError:
        pass
    except Exception:
        return False

    print("[openpyxl_loader] Installing openpyxl in CPython...")
    try:
        subprocess.check_call([python_exe, "-m", "pip", "install", "openpyxl", "--user", "--quiet"])
        print("[openpyxl_loader] openpyxl installed successfully.")
        return True
    except Exception as exc:
        print("[openpyxl_loader] openpyxl installation failed: {}".format(exc))
        return False


def try_import_via_sys_path(python_exe):
    added_paths = []
    for site_path in get_cpython_site_packages(python_exe):
        if os.path.isdir(site_path) and site_path not in sys.path:
            sys.path.insert(0, site_path)
            added_paths.append(site_path)

    try:
        import openpyxl
        print("[openpyxl_loader] Loaded openpyxl {} from CPython site-packages.".format(openpyxl.__version__))
        return openpyxl
    except ImportError as exc:
        for site_path in added_paths:
            if site_path in sys.path:
                sys.path.remove(site_path)
        print("[openpyxl_loader] Direct openpyxl import failed: {}".format(exc))
        return None


_BRIDGE_SCRIPT = r"""
import json
import sys
import traceback

def read_xlsx(path):
    import openpyxl
    workbook = openpyxl.load_workbook(path, data_only=True)
    result = {}
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        result[sheet_name] = []
        for row in worksheet.iter_rows(values_only=True):
            result[sheet_name].append([str(value) if value is not None else "" for value in row])
    workbook.close()
    return result

def write_xlsx(path, sheets_data):
    import openpyxl
    workbook = openpyxl.load_workbook(path)
    for sheet_info in sheets_data:
        sheet_name = sheet_info["name"]
        rows = sheet_info["rows"]
        if sheet_name not in workbook.sheetnames:
            continue
        worksheet = workbook[sheet_name]
        for row_idx, row_data in enumerate(rows):
            for col_idx, value in enumerate(row_data):
                worksheet.cell(row=row_idx + 1, column=col_idx + 1).value = value if value != "" else None
    workbook.save(path)
    workbook.close()

try:
    payload = json.loads(sys.stdin.read())
    if payload["action"] == "read":
        print(json.dumps({"ok": True, "data": read_xlsx(payload["path"])}))
    elif payload["action"] == "write":
        write_xlsx(payload["path"], payload["sheets"])
        print(json.dumps({"ok": True}))
    else:
        print(json.dumps({"ok": False, "error": "Unknown action: " + payload["action"]}))
except Exception:
    print(json.dumps({"ok": False, "error": traceback.format_exc()}))
"""


class SubprocessExcelBridge(object):
    def __init__(self, python_exe):
        self.python_exe = python_exe
        self._bridge_path = self._write_bridge_script()

    def _write_bridge_script(self):
        fd, path = tempfile.mkstemp(suffix=".py", prefix="p13_excel_bridge_")
        with os.fdopen(fd, "w") as bridge_file:
            bridge_file.write(_BRIDGE_SCRIPT)
        return path

    def _call(self, payload):
        proc = subprocess.Popen(
            [self.python_exe, self._bridge_path],
            stdin=subprocess.PIPE,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE
        )
        stdout, stderr = proc.communicate(input=json.dumps(payload).encode("utf-8"))
        
        stdout_str = _decode_output(stdout).strip()
        result = None
        try:
            result = json.loads(stdout_str)
        except ValueError:
            # Try to find a line starting with {"ok":
            json_line = None
            for line in stdout_str.splitlines():
                if line.strip().startswith('{"ok":'):
                    json_line = line.strip()
                    break
            if json_line:
                try:
                    result = json.loads(json_line)
                except ValueError:
                    pass
            
            if result is None:
                raise RuntimeError("Excel bridge did not return JSON.\nstdout={}\nstderr={}".format(stdout_str, _decode_output(stderr)))

        if not result.get("ok"):
            raise RuntimeError("Excel bridge error:\n{}".format(result.get("error", "Unknown error")))
        return result

    def read_xlsx(self, path):
        return self._call({"action": "read", "path": path})["data"]

    def write_xlsx(self, path, sheets_data):
        self._call({"action": "write", "path": path, "sheets": sheets_data})

    def cleanup(self):
        try:
            os.remove(self._bridge_path)
        except Exception:
            pass


def load_openpyxl(auto_install=True):
    try:
        import openpyxl
        return openpyxl
    except ImportError:
        pass

    python_exe = find_cpython()
    if python_exe is None:
        raise RuntimeError("CPython was not found. Install Python from https://python.org or use XLSX native fallback.")

    print("[openpyxl_loader] Found CPython: {}".format(python_exe))
    if auto_install:
        ensure_openpyxl_installed(python_exe)

    openpyxl = try_import_via_sys_path(python_exe)
    if openpyxl is not None:
        return openpyxl

    raise RuntimeError("Direct openpyxl import failed. Use get_subprocess_bridge() instead.")


def get_subprocess_bridge(auto_install=True):
    python_exe = find_cpython()
    if python_exe is None:
        raise RuntimeError("CPython was not found. Install Python from https://python.org or use XLSX native fallback.")

    if auto_install:
        if not ensure_openpyxl_installed(python_exe):
            raise RuntimeError("openpyxl is not installed and auto-installation failed.")

    return SubprocessExcelBridge(python_exe)
